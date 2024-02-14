import pandas as pd
from geopy.distance import geodesic
from fuzzywuzzy import process, fuzz
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, PhotoImage
import threading
import time
from openpyxl.styles import PatternFill, Font
import numpy as np


class ResortMatcherApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Resort Matcher")
        self.root.geometry("1024x900")

        logo_image = PhotoImage(file="logo.png")
        logo_label = tk.Label(self.root, image=logo_image)
        logo_label.image = logo_image
        logo_label.pack(pady=10)

        ttk.Button(self.root, text="Select Excel File", command=self.load_excel).pack(
            pady=10
        )
        self.save_results_button = ttk.Button(
            self.root, text="Save Results", command=self.save_results
        )
        self.save_results_button.pack_forget()

        self.progress = ttk.Progressbar(
            self.root, orient=tk.HORIZONTAL, length=200, mode="determinate"
        )
        self.progress.pack(pady=20)

        self.time_remaining_label = ttk.Label(
            self.root, text="Estimated Time Remaining: N/A"
        )
        self.time_remaining_label.pack(pady=10)

        self.processed_rows_label = ttk.Label(self.root, text="Rows Processed: 0 / 0")
        self.processed_rows_label.pack(pady=10)

        self.cancel_process = False
        self.cancel_button = ttk.Button(
            self.root, text="Cancel", command=self.cancel_operation
        )
        self.cancel_button.pack(pady=10)

        # Setup for the Treeview
        self.tree = ttk.Treeview(
            root, columns=("Name", "Match Status", "Match Score"), show="headings"
        )
        self.tree.heading("Name", text="Resort Name")
        self.tree.heading("Match Status", text="Match Status")
        self.tree.heading("Match Score", text="Match Score")
        self.tree.column("Name", width=150)
        self.tree.column("Match Status", width=100)
        self.tree.column("Match Score", width=30)
        # First, pack the TreeView
        self.tree.pack(expand=True, fill="both", side="top")

        # Then, create and pack the scrollbar
        scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.match_status_frame = ttk.Frame(root)
        self.match_status_frame.pack(fill="x", side="bottom")
        self.matched_resorts_label = ttk.Label(
            self.match_status_frame, text="Matched Resorts: 0"
        )
        self.matched_resorts_label.pack(side="left")
        self.potential_matches_label = ttk.Label(
            self.match_status_frame, text="Potential Matches: 0"
        )
        self.potential_matches_label.pack(side="left")
        self.no_match_found_label = ttk.Label(
            self.match_status_frame, text="No Match Found: 0"
        )
        self.no_match_found_label.pack(side="left")
        self.total_processed_label = ttk.Label(
            self.match_status_frame, text="Total Processed: 0 / 0"
        )
        self.total_processed_label.pack(side="left")

        self.filename = ""
        self.results = None
        self.df_our_resorts = None
        self.df_to_match = None

    def load_excel(self):
        self.filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if self.filename:
            self.progress["value"] = 0
            self.root.update_idletasks()
            threading.Thread(target=self.process_file, daemon=True).start()

    def process_file(self):
        matched_resorts = 0
        potential_matches_count = 0
        no_match_found = 0
        total_steps = 0
        match_results = []

        try:
            self.df_our_resorts = pd.read_excel(self.filename, sheet_name=0)
            self.df_to_match = pd.read_excel(self.filename, sheet_name=1)
            total_steps = len(self.df_our_resorts.index)

            start_time = time.time()

            self.root.after(
                0,
                lambda: self.total_processed_label.config(
                    text=f"Total Processed: 0 / {total_steps}"
                ),
            )
            self.update_progress(10)

            for i, (_, resort_data) in enumerate(
                self.df_our_resorts.iterrows(), start=1
            ):
                if self.cancel_process:
                    return

                match_result = self.find_best_match(
                    resort_data["Resort Name"],
                    resort_data["Resort ID"],
                    resort_data["Resort Street"],
                    resort_data["Resort City"],
                    resort_data["Resort State"],
                    resort_data["Resort Zip"],
                    resort_data.get("Latitude", "N/A"),
                    resort_data.get("Longitude", "N/A"),
                    self.df_to_match,
                )

                elapsed_time = time.time() - start_time
                avg_time_per_item = elapsed_time / i if i > 0 else 0
                estimated_time_remaining = (
                    avg_time_per_item * (total_steps - i) if i < total_steps else 0
                )

                # Update progress including estimated time remaining
                self.update_progress(
                    20 + (80 * i // total_steps), estimated_time_remaining
                )
                self.processed_rows_label.config(
                    text=f"Rows Processed: {i}/{total_steps}"
                )

                if match_result:
                    match_status = match_result[0]["Match Status"]
                    match_score = match_result[0][
                        "Match Score"
                    ]  # Assuming 'Match Score' is calculated in 'find_best_match'
                    if match_status == "Matched Resort":
                        matched_resorts += 1
                        tree_value = (
                            resort_data["Resort Name"],
                            "Matched Resort",
                            match_score,
                        )
                    elif match_status == "Potential Match":
                        potential_matches_count += (
                            1  # Increment potential matches count
                        )
                        tree_value = (
                            resort_data["Resort Name"],
                            "Potential Match",
                            match_score,
                        )
                    else:
                        no_match_found += 1
                        tree_value = (
                            resort_data["Resort Name"],
                            "No Match Found",
                            match_score,
                        )
                else:
                    no_match_found += 1
                    tree_value = (
                        resort_data["Resort Name"],
                        "No Match Found",
                        "N/A",
                    )  # Use "N/A" or a suitable placeholder for no score

                self.tree.insert("", "end", values=tree_value)

                self.update_progress(20 + (80 * i // total_steps), None)
                self.processed_rows_label.config(
                    text=f"Rows Processed: {i}/{total_steps}"
                )

                self.matched_resorts_label.config(
                    text=f"Matched Resorts: {matched_resorts}"
                )
                self.potential_matches_label.config(
                    text=f"Potential Matches: {potential_matches_count}"
                )
                self.no_match_found_label.config(
                    text=f"No Match Found: {no_match_found}"
                )
                self.total_processed_label.config(
                    text=f"Total Processed: {i} / {total_steps}"
                )

                # Collect all match results
                match_results.append(match_result)

            # Here, ensure match_results is correctly structured for pd.concat
            self.results = pd.concat(
                [
                    self.df_our_resorts.reset_index(drop=True),
                    pd.DataFrame(match_results),
                ],
                axis=1,
            )

            flattened_matches = []
            for match_set in match_results:
                # Flatten each set of matches for a single resort entry
                flat_match = {}
                for i, match in enumerate(match_set, start=1):
                    for key, value in match.items():
                        flat_match[f"{key} {i}"] = value
                flattened_matches.append(flat_match)

            # Combine the original DataFrame with the flattened match details
            matches_df = pd.DataFrame(flattened_matches)
            self.final_results = matches_df

            if not self.cancel_process:

                self.update_progress(100)
                self.save_results_button.pack(pady=10)
                messagebox.showinfo(
                    "Process Complete",
                    "File has been processed. You can now save the results.",
                )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process file: {e}")

    def cancel_operation(self):
        self.cancel_process = True
        self.tree.delete(*self.tree.get_children())
        self.reset_gui()

    def reset_gui(self):
        self.progress["value"] = 0
        self.time_remaining_label.config(text="")
        self.processed_rows_label.config(text="Rows Processed: 0 / 0")
        self.matched_resorts_label.config(text="Matched Resorts: 0")
        self.potential_matches_label.config(text="Potential Matches: 0")
        self.no_match_found_label.config(text="No Match Found: 0")
        self.total_processed_label.config(text="Total Processed: 0 / 0")
        messagebox.showinfo(
            "Operation Cancelled", "The operation was cancelled successfully."
        )
        self.cancel_process = False

    def update_progress(self, value, estimated_time_remaining=None):
        self.progress["value"] = value
        self.root.update_idletasks()
        if estimated_time_remaining:
            mins, secs = divmod(estimated_time_remaining, 60)
            time_str = f"Estimated Time Remaining: {int(mins)}m {int(secs)}s"
            self.time_remaining_label.config(text=time_str)

    def find_best_match(
        self, name, id, street, city, state, zip_code, latitude, longitude, df_to_match
    ):

        name_weight = 0.5  # Adjust these weights as necessary
        address_weight = 0.25
        distance_weight = 0.25

        # Adjust weights if the address is incomplete or has NaN values
        if any(
            pd.isna(x) or x in [None, "", " "] for x in [street, city, state, zip_code]
        ):
            name_weight = 0.6
            address_weight = 0  # Set to 0 because address will not be used in scoring
            distance_weight = 0.4

        max_name_similarity = 100
        max_address_similarity = 100
        max_distance = 5  # Define what you consider the maximum relevant distance

        potential_matches = []

        for _, row in df_to_match.iterrows():
            name_similarity = fuzz.ratio(name.lower(), row["Resort Name"].lower())

            # Calculate address similarity only if all parts of the address are provided and not NaN
            address_similarity = 0
            if all([street, city, state, zip_code]) and not any(
                pd.isna(x) for x in [street, city, state, zip_code]
            ):
                full_address_input = f"{street}, {city}, {state}, {zip_code}".lower()
                full_address_row = f"{row['Resort Street']}, {row['Resort City']}, {row['Resort State']}, {row['Resort Zip']}".lower()
                address_similarity = fuzz.ratio(full_address_input, full_address_row)

            # Calculate distance if latitudes and longitudes are valid
            distance = float("inf")
            if not any(
                np.isnan(x)
                for x in [latitude, longitude, row["Latitude"], row["Longitude"]]
            ):
                distance = geodesic(
                    (latitude, longitude), (row["Latitude"], row["Longitude"])
                ).miles

            # Normalize the metrics to a scale of 0 to 100
            normalized_name_similarity = (name_similarity / max_name_similarity) * 100
            normalized_address_similarity = (
                (address_similarity / max_address_similarity) * 100
                if address_similarity is not None
                else 0
            )
            normalized_distance = (
                1 - min(distance / max_distance, 1)
            ) * 100  # Inverted because a smaller distance is better

            # Calculate the weighted score for each metric
            weighted_name_score = normalized_name_similarity * name_weight
            weighted_address_score = normalized_address_similarity * address_weight
            weighted_distance_score = normalized_distance * distance_weight

            # Calculate the final combined score
            combined_score = (
                weighted_name_score + weighted_address_score + weighted_distance_score
            )

            # Determine the match status based on the combined score
            match_status = ""
            if combined_score >= 80:
                match_status = "Matched Resort"
            elif 50 <= combined_score < 80:
                match_status = "Potential Match"
            else:
                match_status = "No Match Found"

            potential_matches.append(
                {
                    "Match Status": match_status,
                    "Match Score": combined_score,
                    "Resort Name": name,
                    "Resort ID": id,
                    "Matching Resort Name": row["Resort Name"],
                    "Matching Resort ID": row["Resort ID"],
                    "Name Similarity": name_similarity,
                    "Address Similarity": address_similarity,
                    "Distance (miles)": distance,
                }
            )

        # Sort matches based on probability, then by distance and similarities
        potential_matches.sort(key=lambda x: x["Match Score"], reverse=True)

        # Filter and return matches based on status
        matched_resorts = [
            match
            for match in potential_matches
            if match["Match Status"] == "Matched Resort"
        ]
        return matched_resorts[:1] if matched_resorts else potential_matches[:5]

    def save_results(self):
        if self.final_results is not None:
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if save_path:
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    self.final_results.to_excel(
                        writer, sheet_name="Processed Results", index=False
                    )
                    self.df_our_resorts.to_excel(
                        writer, sheet_name="Our Resorts", index=False
                    )
                    self.df_to_match.to_excel(
                        writer, sheet_name="Resorts to Match", index=False
                    )

                    # Additional sheets if needed
                    workbook = writer.book
                    worksheet = writer.sheets["Processed Results"]

                    # Apply formatting
                    red_fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid"
                    )
                    green_fill = PatternFill(
                        start_color="00B050", end_color="00B050", fill_type="solid"
                    )
                    yellow_fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )
                    white_font = Font(color="FFFFFF")

                    for row in worksheet.iter_rows(
                        min_row=2, max_col=1, max_row=worksheet.max_row
                    ):
                        for cell in row:
                            if cell.value == "Matched Resort":
                                cell.fill = green_fill
                                cell.font = white_font
                            elif cell.value == "Potential Match":
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill
                                cell.font = white_font

                    # Save the workbook to apply the changes
                    workbook.save(save_path)

                messagebox.showinfo("Save Complete", "Results have been saved.")
        else:
            messagebox.showwarning("No Data", "Please process a file first.")


if __name__ == "__main__":
    root = tk.Tk()
    app = ResortMatcherApp(root)
    root.mainloop()
